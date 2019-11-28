#=============================
# Name       :format.py
# Argument   :
# Email      :lukyandy3162@gmail.com
# Author     :srhuang
# History    :
#    20191127:Initial
#=============================

#===============
#import section
#===============
import os
import sys
import openpyxl
from openpyxl.styles import Font


#================
#variable section
#================
workbook='example'
input='input/'+workbook+'.xlsx'
output='output/'+workbook+'.xlsx'
target_sheet='sheet1'

#=================
#argument section
#=================

#=================
#function section
#=================

#===============
#progress start
#===============

# open work book
wb=openpyxl.load_workbook(input)
sheet=wb.get_sheet_by_name(target_sheet)

# get Font object
# name / size / bold / italic
Font1 = Font(name='Calibri', size=24, italic=False, bold=True)
sheet['A1'].font=Font1

# adjust column/row height and width
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 70

# merge cells
sheet.merge_cells('C5:D5')

#unmerge cells
sheet.unmerge_cells('C5:D5')

#freeze panes
# 'A2' : row 1
# 'B1' : column A
# 'C1' : column A and B
# 'C2' : row 1, column A, column B
# 'A1' : no freeze
# None : no freeze
sheet.freeze_panes='A2'

#save
wb.save(output)




