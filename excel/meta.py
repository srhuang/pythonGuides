#=============================
# Name       :meta.py
# Argument   :
# Email      :lukyandy3162@gmail.com
# Author     :srhuang
# History    :
#    20191128:Initial
#=============================

#===============
#import section
#===============
import os
import sys
import openpyxl
#import pandas as pd


#================
#variable section
#================
workbook='input/example.xlsx'
target_sheet='sheet1'

#=================
#argument section
#=================

#=================
#function section
#=================

#===============
#progress start
#===============

#open work book
wb=openpyxl.load_workbook(workbook)

#get sheets
sheet=wb.get_sheet_by_name(target_sheet)

#get max column and max row
print "max row :"+str(sheet.max_row)
print "max column :"+str(sheet.max_column)
