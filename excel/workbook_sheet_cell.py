#=============================
# Name       :workbook_sheet_cell.py
# Argument   :
# Email      :lukyandy3162@gmail.com
# Author     :srhuang
# History    :
#    20191126:Initial
#=============================

#===============
#import section
#===============
import os
import sys
import openpyxl

#================
#variable section
#================
workbook='input/example.xlsx'
target_sheet='sheet1'

#=================
#argument section
#=================

#=================
#function section
#=================

#===============
#progress start
#===============

#open work book
wb=openpyxl.load_workbook(workbook)

#get sheets
print "-----All Sheets-----"
all_sheet=wb.get_sheet_names()
for sheet in all_sheet:
    if sheet==target_sheet:
        sheet_name=sheet
    print sheet

print "-----Get Sheet-----"
sheet=wb.get_sheet_by_name(sheet_name)
print sheet['A1'].value

#get cell
print "-----Get Cell-----"
print sheet.cell(row=1, column=2).value

#update sheet
all_sheet=wb.get_sheet_names()
sheet2=wb.get_sheet_by_name(all_sheet[1])
sheet2.title='new'
wb.save(workbook)

#new and delete sheet
wb.create_sheet(index=3, title='new')
wb.save(workbook)

all_sheet=wb.get_sheet_names()
wb.remove_sheet(wb.get_sheet_by_name(all_sheet[2]))
print "-----Update the Sheets-----"
print wb.get_sheet_names()
wb.save(workbook)

