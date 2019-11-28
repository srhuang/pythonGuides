#=============================
# Name       :read_data.py
# Argument   :
# Email      :lukyandy3162@gmail.com
# Author     :srhuang
# History    :
#    20191128:Initial
#=============================

#===============
#import section
#===============
import os
import sys
import openpyxl

#================
#variable section
#================
workbook='input/example.xlsx'
target_sheet='sheet1'

#=================
#argument section
#=================

#=================
#function section
#=================

#===============
#progress start
#===============

#open work book
wb=openpyxl.load_workbook(workbook)

#get sheets
sheet=wb.get_sheet_by_name(target_sheet)

#get cell
print "-----Get Cell-----"
print sheet['A1'].value
print sheet.cell(row=1, column=2).value

#1~11, add 2 for each step
print "-----Print Each Two Row-----"
for i in range(2,13,2):
    print(i, sheet.cell(row=i, column=2).value)

#get area
print "-----Get Area-----"
for rowOfCellObject in sheet['A2':'C9']:
    #print each row
    for cellObject in rowOfCellObject:
        print cellObject.coordinate, cellObject.value 
    print "---END OF ROW---"

#get specific column
print "-----Get Specific Column-----"
for cellObject in sheet['B']:
    print cellObject.value

#get specific row
print "-----Get Specific Row-----"
for cell in sheet['1']:
    print cell.value


